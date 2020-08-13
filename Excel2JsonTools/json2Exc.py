#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import json
from openpyxl import Workbook


def read_json(file):
    """
    打开设置路径下的json，并json中的内容。 （此时的内容为dict 格式）
    """
    with open(file, 'r', encoding='utf8') as fp:
        json_data = json.load(fp)
        return json_data

def write2xlsx(read_path, write_path, sheet_name):
    """
    将json中的内容，写入xlsx文件中。 
    json 的格式与想要转化成excel 的格式都是case by case， 通用很难，所以遇到了，要自己修改逻辑。
    """
    json = read_json(read_path)
    item_list = json['items']
    excel = Workbook()
    sheet = excel.create_sheet(sheet_name, index=0)
    v = 0
    for ig in item_list:
        dic_item = dict(ig)
        if v == 0:
            for i, key in enumerate(dic_item):
                sheet.cell(row=1, column=i+1, value=key)
                sheet.cell(row=v+2, column=i+1, value=dic_item[key])
        if v > 0:
            for i, key in enumerate(dic_item):
                sheet.cell(row=v+2, column=i+1, value=dic_item[key])
        v = v+1
    excel.save(write_path)
    return 0
